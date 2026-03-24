from __future__ import annotations

from pathlib import Path


def extract_text_from_path(path: Path, doc_type: str) -> str:
    """从本地文件抽取纯文本；不支持的类型返回空串。"""
    suffix = path.suffix.lower()
    try:
        if suffix in (".txt", ".md", ".markdown"):
            return path.read_text(encoding="utf-8", errors="ignore")
        if suffix == ".pdf" or doc_type.upper() == "PDF":
            return _pdf(path)
        if suffix in (".docx",) or doc_type.upper() == "WORD":
            return _docx(path)
        if suffix in (".xlsx", ".xls") or doc_type.upper() == "EXCEL":
            return _xlsx(path)
    except Exception:
        return ""
    return ""


def _pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    parts: list[str] = []
    for page in reader.pages:
        try:
            parts.append(page.extract_text() or "")
        except Exception:
            continue
    return "\n".join(parts)


def _docx(path: Path) -> str:
    from docx import Document

    d = Document(str(path))
    return "\n".join(p.text for p in d.paragraphs if p.text)


def _xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                lines.append(" ".join(cells))
    return "\n".join(lines)
